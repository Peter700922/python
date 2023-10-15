import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import requests
from PIL import Image,ImageTk
import io
from urllib.request import urlopen
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors
from openpyxl.styles import PatternFill  
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import threading
import time
from threading import Timer

root = tk.Tk()
root.title("currency exchange")
root.geometry("400x400")
#icon
url = "https://www.icosky.com/icon/ico/Business/Financial/Green%20Dollar.ico"
image_bytes = urlopen(url).read()
# internal data file
data_stream = io.BytesIO(image_bytes)
# open as a PIL image object
pil_image = Image.open(data_stream)
root.iconphoto(False, ImageTk.PhotoImage(pil_image))
#bgimg
url = "https://cdn2.vectorstock.com/i/1000x1000/24/76/piggy-bank-line-icon-finance-and-banking-vector-25912476.jpg"
image_bytes = urlopen(url).read()
# internal data file
data_stream = io.BytesIO(image_bytes)
# open as a PIL image object
pil_image = Image.open(data_stream)
pil_image = pil_image.resize((400, 400))
# convert PIL image object to Tkinter PhotoImage object
tk_image = ImageTk.PhotoImage(pil_image) 

# 建立不斷改變文字變數的函式
def showTime():
    GMT = datetime.timezone(datetime.timedelta(hours=8))    # 設定所在時區 ( 台灣是 GMT+8 )
    now = datetime.datetime.now(tz=GMT).strftime('%Y-%m-%d %H:%M:%S')   # 取得目前的時間，格式使用 H:M:S
    localTime['text'] = str(now)
    localTime['bg'] = 'lightgreen'
    localTime['font']=('Arial',16,'bold')
    root.after(1000, showTime)    # 視窗每隔 1000 毫秒再次執行一次 showTime()
#程式部分
def inputBoxFuc(event):
    inputBox.delete(0, tk.END)
def change():
    if str.isdigit(inputBox.get()) == False:
        messagebox.showwarning(title="錯誤通知", message = '請輸入正確金額')
        return
    if inputSec.current() == -1:
        messagebox.showwarning(title="錯誤通知", message = '請選擇輸入貨幣')
        return
    if changeSec.current() == -1:
        messagebox.showwarning(title="錯誤通知", message = '請選擇更換貨幣')
        return
    moneyChange()
class RepeatingTimer(Timer): 
    def run(self):
        self.finished.wait(self.interval)
        while not self.finished.is_set():
            self.function(*self.args, **self.kwargs)
            self.finished.wait(self.interval)    

def moneyChange():
    currency = ['JPY','GBP', 'CHF', 'CNY', 'BTC', 'ETH']
    sign = ['日元', '英磅', '歐元', '人民幣','比特幣', '乙太幣']
    input = str(currency[inputSec.current()])
    out = currency[changeSec.current()]
    outSign = sign[changeSec.current()]
    API = requests.get("https://api.coinbase.com/v2/exchange-rates?currency="+input)
    exchange = float(API.json()['data']['rates'][out])
    exchangeMoney = round(exchange * float(inputBox.get()), 3)
    status_label['text'] = str(exchangeMoney) + outSign
    status_label['fg'] = '#f00'
    GMT = datetime.timezone(datetime.timedelta(hours=8))    # 設定所在時區 ( 台灣是 GMT+8 )
    now = datetime.datetime.now(tz=GMT).strftime('%Y-%m-%d %H:%M:%S')   # 取得目前的時間，格式使用 H:M:S
    systemTime['text'] = str(now)
    systemTime['bg'] = 'lightblue'
    systemTime['font']=('Arial',16,'bold')

def report():
    book = Workbook()
    sheet = book.active
    #title
    backGround =sheet["A1:Z38"]
    for i in backGround:
        for j in i:
            j.fill = PatternFill(fill_type="solid", fgColor="b7f1db")  
            j.font = Font(
                name='Arial',               # 字体名
                strike=None,                # 删除线，True/False
                color=colors.BLACK,         # 文字颜色
                size=14,                    # 字的大小
                bold=True,                  # 加粗, True/False
                italic=None,                # 倾斜，Tue/False
                underline=None              # 下划线: 'singleAccounting', 'double', 'single', 'doubleAccounting'
                ) 
            
    sheet['C5'] = "幣值交易系統"
    sheet['C5'].font = Font(
        name='Arial',               # 字体名
        strike=None,                # 删除线，True/False
        color='ff4000',             # 文字颜色
        size=30,                    # 字的大小
        bold=True,                  # 加粗, True/False
        italic=None,                # 倾斜，Tue/False
        underline=None              # 下划线: 'singleAccounting', 'double', 'single', 'doubleAccounting'
        )  
    border=Border(left=Side(border_style='medium',color=colors.BLACK),
                right=Side(border_style='medium',color=colors.BLACK),
                top=Side(border_style='medium',color=colors.BLACK), 
                bottom=Side(border_style='medium',color=colors.BLACK), 
                diagonal=Side(border_style='medium',color=colors.BLACK), 
                diagonal_direction=0, 
                outline=Side(border_style='medium',color=colors.BLACK),
                vertical=Side(border_style='medium',color=colors.BLACK), 
                horizontal=Side(border_style='medium',color=colors.BLACK)
                 )

    sheet["C5"].border = border
    sheet.merge_cells("C5:J7")
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet['C5'].alignment = align
    sheet['C5'].fill = PatternFill(fill_type="solid", fgColor="00bfff")
    sheet['C8'] = "順序"
    sheet['E8'] = "匯率"
    sheet['G8'] = "系統時間"
    sheet['I8'] = "客戶"
    

    #rows
    dataRows = sheet['A8:J38']
    for rows in dataRows:
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(fill_type="solid", fgColor="ff9400")  
            cell.border = border
    for i in range(8,39,1):
        row = 'A'+str(i)+':'+'B'+str(i)
        sheet.merge_cells(row)
        row = 'C'+str(i)+':'+'D'+str(i)
        sheet.merge_cells(row)
        row = 'E'+str(i)+':'+'F'+str(i)
        sheet.merge_cells(row)
        row = 'G'+str(i)+':'+'H'+str(i)
        sheet.merge_cells(row)
        row = 'I'+str(i)+':'+'j'+str(i)
        sheet.merge_cells(row)

    sheet.column_dimensions['G'].width = 20
    for i in range(8,39,1):
        sheet.row_dimensions[i].height=20
    currency = ['JPY','GBP', 'CHF', 'CNY', 'BTC', 'ETH']
    u = 9
    y = 1
    for i in currency:
        API = requests.get("https://api.coinbase.com/v2/exchange-rates?currency="+i)
        for j in currency:
            if i == j:
                continue
            exchange = str(round(float(API.json()['data']['rates'][j]),4))
            code = '\u0045'+str(u)
            sheet[code] = exchange
            code = '\u0043'+str(u)
            sheet[code] = y  
            sheet[code].alignment = Alignment(horizontal='left', wrap_text=True)
            code = '\u0047'+str(u)
            sheet.row_dimensions[u].width = 50
            GMT = datetime.timezone(datetime.timedelta(hours=8))    # 設定所在時區 ( 台灣是 GMT+8 )
            now = datetime.datetime.now(tz=GMT).strftime('%Y-%m-%d %H:%M:%S')   # 取得目前的時間，格式使用 H:M:S
            sheet[code] = str(now) 
            u+=1
            y+=1
    dataRows = sheet['I9:J38']
    for i in range(9,39,1):
        sheet.cell(i,9).value = "吳O賢"
    dataRows = sheet['A9:A38']
    y=9
    sign = ['日幣','英鎊', '法郎', '人民幣', '比特幣', '乙太幣']
    for j in range(0,len(sign) ,1):
         for k in range(0,len(sign) ,1):
            if sign[j]==sign[k] :
                continue
            sheet.cell(y,1).value = sign[j]+"對"+sign[k]+"匯率"
            y+=1  
    sheet.column_dimensions['A'].width = 20
    book.save("report.xlsx")
    reportBar['text'] = "成功輸出報表"
    reportBar['bg'] = 'red'
    reportBar['font']=('Arial',14,'bold')

global x 
global y 
x =([])
y =([]) 
seco = 0.0
def lib() : 
    global seco
    f = Figure(figsize=(2.5,2),dpi=85)
    a = f.add_subplot(111)
    x.append(float(seco))
    y.append(80.0)
    a.plot(x,y)
    canvas =FigureCanvasTkAgg(f, master = root )
    canvas.get_tk_widget().place(x=130 ,y =75)
    seco+=1.0
def thread() :
    #另外執行圖形
    a = RepeatingTimer(3.0, lib)
    a.start()
     
#UI部分
# Show image using label 
label1 = tk.Label(root, image = tk_image) 
label1.place(x = 0, y = 0) 

inputSec = ttk.Combobox(root, width = 10,values=['日幣','英鎊', '法郎', '人民幣', '比特幣', '乙太幣'],font=('Arial',12,'bold'))
inputSec.place(x=10, y=10) 

inputBox= tk.Entry(root,width=15,font=('Arial',12,'bold'))
inputBox.insert(0,'請輸入更換金額')
inputBox.place(x=150, y=10)
inputBox.bind('<ButtonPress-1>', inputBoxFuc)

tk.Button(root, text = "更換", command = change,font=('Arial',14,'bold'),bg = 'cyan').place(x=300, y=5) 

changeSec = ttk.Combobox(root, width = 10,values=['日幣','英鎊', '法郎', '人民幣', '比特幣', '乙太幣'],font=('Arial',12,'bold'))
changeSec.place(x=10, y=50) 

status_label = tk.Label(text='更換匯率',bg='lightgray',font=('Arial',14,'bold'))
status_label.place(x=150, y=50) 

localTimeBar = tk.Label(text='當前時間',bg='lightgray',font=('Arial',16,'bold'))
localTimeBar.place(x=10, y=250) 
localTime = tk.Label(text='')
localTime.place(x=120, y=250)
systemTimeBar = tk.Label(text='系統時間',bg='lightgray',font=('Arial',16,'bold'))
systemTimeBar.place(x=10, y=300) 
systemTime = tk.Label(text='')
systemTime.place(x=120, y=300)
tk.Button(root, text = "產生報表", command = report,font=('Arial',14,'bold'),bg = 'lightblue').place(x=10, y=350) 
reportBar = tk.Label(text='',font=('Arial',14,'bold'))
reportBar.place(x=120, y=350)
tk.Button(root, text = "產生圖表", command = thread,font=('Arial',14,'bold'),bg = 'lightyellow').place(x=10, y=80)  
showTime()
root.mainloop()